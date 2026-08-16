from __future__ import annotations

from pathlib import Path

from .extract import DependencyUnavailableError

_MODULE_HEADERS = {"module", "模块"}
_FUNCTION_HEADERS = {"function", "函数", "函数名"}
_COUNT_HEADERS = {"count", "coverage", "coverage count", "覆盖次数", "执行次数"}
_BRANCH_ID_HEADERS = {"branch_id", "branch id", "分支id", "分支编号"}
_CONDITION_HEADERS = {"condition", "branch", "分支条件", "条件"}
_TRUE_COUNT_HEADERS = {"true_count", "true count", "真分支次数"}
_FALSE_COUNT_HEADERS = {"false_count", "false count", "假分支次数"}


def _column(headers: list[str], accepted: set[str]) -> int | None:
    for index, value in enumerate(headers):
        if value.strip().lower() in accepted:
            return index
    return None


def parse_coverage_xlsx(path: Path) -> tuple[list[dict], list[str]]:
    """Read function and branch execution counts from coverage workbooks."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DependencyUnavailableError("openpyxl", "coverage XLSX") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    warnings: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            first = next(rows, None)
            if first is None:
                continue
            headers = ["" if value is None else str(value) for value in first]
            module_index = _column(headers, _MODULE_HEADERS)
            function_index = _column(headers, _FUNCTION_HEADERS)
            count_index = _column(headers, _COUNT_HEADERS)
            if None not in (module_index, function_index, count_index):
                for row_number, row in enumerate(rows, 2):
                    module = row[module_index] if module_index < len(row) else None
                    function = row[function_index] if function_index < len(row) else None
                    count = row[count_index] if count_index < len(row) else None
                    if module is None and function is None and count is None:
                        continue
                    try:
                        numeric_count = int(count)
                    except (TypeError, ValueError):
                        warnings.append(f"sheet {sheet.title} row {row_number}: invalid coverage count {count!r}")
                        continue
                    records.append({
                        "coverage_type": "function",
                        "module": "" if module is None else str(module),
                        "function": "" if function is None else str(function),
                        "count": numeric_count,
                        "source": str(path),
                        "sheet": sheet.title,
                        "row": row_number,
                    })
                continue

            branch_id_index = _column(headers, _BRANCH_ID_HEADERS)
            condition_index = _column(headers, _CONDITION_HEADERS)
            true_count_index = _column(headers, _TRUE_COUNT_HEADERS)
            false_count_index = _column(headers, _FALSE_COUNT_HEADERS)
            if None in (branch_id_index, function_index, condition_index, true_count_index, false_count_index):
                continue
            for row_number, row in enumerate(rows, 2):
                branch_id = row[branch_id_index] if branch_id_index < len(row) else None
                function = row[function_index] if function_index < len(row) else None
                condition = row[condition_index] if condition_index < len(row) else None
                true_count = row[true_count_index] if true_count_index < len(row) else None
                false_count = row[false_count_index] if false_count_index < len(row) else None
                if all(value is None for value in (branch_id, function, condition, true_count, false_count)):
                    continue
                try:
                    numeric_true = int(true_count)
                    numeric_false = int(false_count)
                except (TypeError, ValueError):
                    warnings.append(
                        f"sheet {sheet.title} row {row_number}: invalid branch counts "
                        f"{true_count!r}/{false_count!r}"
                    )
                    continue
                records.append({
                    "coverage_type": "branch",
                    "branch_id": "" if branch_id is None else str(branch_id),
                    "module": "",
                    "function": "" if function is None else str(function),
                    "condition": "" if condition is None else str(condition),
                    "true_count": numeric_true,
                    "false_count": numeric_false,
                    "count": numeric_true + numeric_false,
                    "source": str(path),
                    "sheet": sheet.title,
                    "row": row_number,
                })
    finally:
        workbook.close()
    return records, warnings


def match_coverage_records(records: list[dict], inventory: dict) -> dict:
    """Match coverage to in-scope symbols without treating execution as risk coverage."""
    symbols: dict[str, list[dict]] = {}
    for file in inventory.get("files", []):
        for function in file.get("functions", []):
            symbols.setdefault(function["symbol"], []).append({
                "repo_id": file["repo_id"],
                "path": file["path"],
                "line": function["line"],
            })
    matched: list[dict] = []
    unmatched: list[dict] = []
    ambiguous: list[dict] = []
    for record in records:
        candidates = symbols.get(record["function"], [])
        meaning = (
            "branch_execution_reference_only"
            if record.get("coverage_type") == "branch"
            else "function_execution_reference_only"
        )
        item = {**record, "matches": candidates, "meaning": meaning}
        if len(candidates) == 1:
            matched.append(item)
        elif candidates:
            ambiguous.append(item)
        else:
            unmatched.append(item)
    return {"matched": matched, "ambiguous": ambiguous, "unmatched": unmatched}
